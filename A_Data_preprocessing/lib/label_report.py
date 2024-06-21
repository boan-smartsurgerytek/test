import time
import copy
import json
import pandas as pd
import xlwings as xw
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill,Font
from os import listdir
from os.path import isfile, join, splitext
import numpy

class CreateSummaryExcel(object):
    
    def __init__(self, root_folder_path, save_path, detect_label_list):
        self.root_folder_path = root_folder_path
        self.save_path = save_path
        self.detect_label_list = detect_label_list
        
        self.label_dict = self.create_label_dict()
        self.col_titles = self.create_col_titles()
        self.summary_col_titles = self.create_summary_col_titles()
        self.summary_data_frame = pd.DataFrame()
        self.run()
        
    def run(self):
        folder_name_list = listdir(self.root_folder_path)
        folder_name_list.sort(reverse=True)
        sheet_name1 = 'Sheet'   #創建EXCEL的多於空表單
        self.wb = Workbook() 
        # self.wb = xw.Book()
        
        # self.col_titles         每個資料夾抬頭
        # self.summary_col_titles 總表抬頭
        rowCount=0
        for folder_name in folder_name_list:
            time_start = time.time()
            self.folder_path = join(self.root_folder_path, folder_name)
            self.count = 1
            self.sht = self.wb.create_sheet(folder_name)
            # self.sht.cell(1,1).value = self.col_titles
            self.sht.append(self.col_titles)
            file_list = self.create_file_list()
            self.write_file_information(file_list)
            self.calculate_video_summary(folder_name)
            time_end = time.time()
            rowCount += 1
            if (rowCount%100) == 0:                          #--------------------------new20221108-yita----------------
                print("處理進度DataNum:" + str(rowCount) )    #--------------------------new20221108-yita----------------   
            print("Summary {} cost {} sec".format(folder_name, round(time_end - time_start, 2)))
            
        self.create_all_count_summary()
        print(self.wb.sheetnames)
        worksheet = self.wb[sheet_name1]
        self.wb.remove(worksheet)
        print(self.wb.sheetnames)
        self.wb.save(self.save_path)
        # self.create_all_count_summary()
            
    def create_label_dict(self):
        label_dict = dict()
        count = 1
        for label in self.detect_label_list:
            label_dict[label] = count
            count += 1
        return label_dict
            
    def create_col_titles(self):
        col_titles = ["Image Name"]
        col_titles.extend(self.detect_label_list)
        col_titles.extend(["", "Number of label categories"])
        return col_titles
            
    def create_summary_col_titles(self):
        summary_col_titles = copy.copy(self.col_titles)
        summary_col_titles[0] = "Folder Name"
        summary_col_titles[-1] = "Number of Image"
        return summary_col_titles
    
    def create_file_list(self):
        file_list = list({x[:x.find(".")] for x in listdir(self.folder_path)})
        file_list.sort()
        if "Thumbs" in file_list:
            file_list.remove("Thumbs")
        return file_list
    
    def write_file_information(self, file_list):
        file_list = sorted(file_list)#, key=lambda x: int(splitext(x)[0].split("_")[1]))
        for file_name in file_list:
            json_path = join(self.folder_path, file_name+'.json')
            if isfile(json_path):
                with open(json_path, 'r') as f:
                    file_label_dict = json.load(f)
                    
                fi = self.create_file_information(file_name, file_label_dict)

                self.count += 1
                # self.sht.cell(1,self.count).value = fi
                self.sht.append(fi)
                # self.sht.range('A'+str(self.count)).value = fi
                
    def create_file_information(self, file_name, file_label_dict):
        file_information = [0] * len(self.col_titles)
        file_information[0] = file_name
        
        dict_value_list = list(file_label_dict.values())
        dict_value_set = set(file_label_dict.values())
        dict_value_set = dict_value_set.intersection(set(self.detect_label_list))
        
        for class_name in dict_value_set:
            file_information[self.label_dict[class_name]] += dict_value_list.count(class_name)
            
        file_information[-1] = sum([x != 0 for x in file_information[1:-2]])
        write_file_information = ['' if x == 0 else x for x in file_information]
        return write_file_information
    
    def calculate_video_summary(self, folder_name):
        video_summary = [folder_name]
        last_col_str = self.trans_num_to_uppercase_eng(len(self.col_titles))
        cellend=last_col_str+str(self.count)
        caseVal = {} 
        case=[]
        for index, item in enumerate(self.sht["B2":cellend]):
            # print(index)
            case=[]
            for cell in item:
                if cell.value=="":
                    cell.value=numpy.nan
                    # print(cell.value)
                case.append(cell.value)
            caseVal[index]=case
        # print(caseVal)
                
        data = pd.DataFrame(caseVal)
        data = data.T
        
        
        # data = pd.DataFrame(self.sht.cell("B2:"+last_col_str+str(self.count)).value)
        count_summary = list(data.count())
        count_summary[-2] = ""
        video_summary.extend(count_summary)
        self.summary_data_frame = self.summary_data_frame.append([video_summary])
        # print(data)
        # print(list(data.count()))
        
    def trans_num_to_uppercase_eng(self, num):
        string = ""
        while num > 0:
            num, remainder = divmod(num - 1, 26)
            string = chr(65 + remainder) + string
        return string
    
    def create_all_count_summary(self):
        # self.summary_data_frame.columns = self.summary_col_titles
        # self.summary_data_frame.rows = self.summary_col_titles
        # self.summary_data_frame.columns = self.summary_col_titles
        self.sht = self.wb.create_sheet("Summary",0)
        
        # [h, l] = self.summary_data_frame.shape  # h为行数，l为列数
        self.sht.append(self.summary_col_titles)
        myArr = numpy.array(self.summary_data_frame)
        for x in myArr:
            self.sht.append(x.tolist())
            # print(x.tolist())

        # self.sht.cell(1,1).options(index=False).value = self.summary_data_frame
        
        